from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
import openpyxl
import pandas as pd
import re
import os
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime
from pathlib import Path

app = FastAPI(title="Excel Regex Pro API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3005",
        "http://localhost:3000",
        "https://excelprov1.vercel.app",
        "https://excelprov1-1.onrender.com",
        "https://*.vercel.app",
        "*"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "uploads"
OUTPUT_DIR = "outputs"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Serve React build files
FRONTEND_BUILD = Path(__file__).parent.parent / "frontend" / "build"
if FRONTEND_BUILD.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND_BUILD / "static")), name="static")

class FindReplaceRequest(BaseModel):
    filename: str
    pattern: str
    replacement: str
    sheet_name: Optional[str] = None
    case_sensitive: bool = True

class SheetMatchRequest(BaseModel):
    filename: str
    pattern: str

class VLookupRequest(BaseModel):
    filename: str
    lookup_sheet: str
    lookup_column: str
    lookup_value: Optional[str] = None
    lookup_value_pattern: Optional[str] = None
    target_sheet: str
    target_column: str
    result_column: str
    match_type: str = "exact"  # exact, contains, regex

class ExtractDataRequest(BaseModel):
    filename: str
    pattern: str
    sheet_name: Optional[str] = None
    capture_groups: bool = True

class ValidateDataRequest(BaseModel):
    filename: str
    pattern: str
    column_index: int
    sheet_name: Optional[str] = None

class RegexVLOOKUPRequest(BaseModel):
    filename: str
    lookup_pattern: str
    lookup_sheet: str
    lookup_column: int
    return_column: int
    target_sheet: str
    target_column: int
    case_sensitive: bool = True

class ConditionalFormatRequest(BaseModel):
    filename: str
    pattern: str
    sheet_name: str
    column_index: Optional[int] = None
    highlight_color: str = "FFFF00"

class SplitColumnRequest(BaseModel):
    filename: str
    sheet_name: str
    column_index: int
    split_pattern: str
    max_splits: Optional[int] = None

class MergeColumnsRequest(BaseModel):
    filename: str
    sheet_name: str
    column_indices: List[int]
    pattern: str
    separator: str = " "

class MergeSheetsRequest(BaseModel):
    filename: str
    left_sheet: str
    right_sheet: str
    left_key: str
    right_key: str
    merge_type: str = "inner"  # inner, left, right, outer

class XLookupRequest(BaseModel):
    filename: str
    lookup_sheet: str
    lookup_column: str
    target_sheet: str
    target_column: str
    return_columns: List[str]
    match_type: str = "exact"

class ReconcileRequest(BaseModel):
    filename: str
    sheet1: str
    sheet2: str
    key_column: str
    compare_columns: List[str]

@app.get("/")
async def root():
    return {"message": "Excel Regex Pro API"}

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload Excel file"""
    try:
        file_path = os.path.join(UPLOAD_DIR, file.filename)
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        workbook = openpyxl.load_workbook(file_path)
        sheets = workbook.sheetnames
        
        return {
            "filename": file.filename,
            "sheets": sheets,
            "message": "File uploaded successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/find-replace")
async def find_replace(request: FindReplaceRequest):
    """Find and replace with regex pattern"""
    try:
        file_path = os.path.join(UPLOAD_DIR, request.filename)
        workbook = openpyxl.load_workbook(file_path)
        
        flags = 0 if request.case_sensitive else re.IGNORECASE
        matches_count = 0
        
        sheets_to_process = [request.sheet_name] if request.sheet_name else workbook.sheetnames
        
        for sheet_name in sheets_to_process:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        new_value = re.sub(request.pattern, request.replacement, cell.value, flags=flags)
                        if new_value != cell.value:
                            matches_count += 1
                            cell.value = new_value
        
        output_filename = f"modified_{request.filename}"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        workbook.save(output_path)
        
        return {
            "matches": matches_count,
            "output_file": output_filename,
            "message": f"Replaced {matches_count} matches"
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/sheet-match")
async def sheet_match(request: SheetMatchRequest):
    """Find sheets matching pattern"""
    try:
        file_path = os.path.join(UPLOAD_DIR, request.filename)
        workbook = openpyxl.load_workbook(file_path)
        
        matched_sheets = [
            sheet for sheet in workbook.sheetnames 
            if re.search(request.pattern, sheet, re.IGNORECASE)
        ]
        
        return {
            "matched_sheets": matched_sheets,
            "total_matches": len(matched_sheets)
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/sheet-columns")
async def get_sheet_columns_by_name(filename: str, sheet_name: str = None):
    """Get columns for a specific sheet or all sheets"""
    try:
        file_path = os.path.join(UPLOAD_DIR, filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"File not found: {filename}")
        
        if sheet_name:
            # Get columns for specific sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            return {"columns": list(df.columns)}
        else:
            # Get columns for all sheets
            df_workbook = pd.read_excel(file_path, sheet_name=None)
            columns = {}
            for sname, df in df_workbook.items():
                columns[sname] = list(df.columns)
            return {"columns": columns}
    except Exception as e:
        print(f"Error in get_sheet_columns: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/vlookup")
async def vlookup(request: VLookupRequest):
    """Perform VLOOKUP operation between sheets"""
    try:
        file_path = os.path.join(UPLOAD_DIR, request.filename)
        df_workbook = pd.read_excel(file_path, sheet_name=None)
        
        if request.lookup_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.lookup_sheet}' not found")
        if request.target_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.target_sheet}' not found")
        
        lookup_df = df_workbook[request.lookup_sheet]
        target_df = df_workbook[request.target_sheet]
        
        # Check if columns exist
        if request.lookup_column not in lookup_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.lookup_column}' not found in lookup sheet")
        if request.target_column not in target_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.target_column}' not found in target sheet")
        if request.result_column not in target_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.result_column}' not found in target sheet")
        
        matches_count = 0
        results = []
        
        # Perform lookup
        for idx, row in lookup_df.iterrows():
            lookup_val = str(row[request.lookup_column])
            
            if request.match_type == "exact":
                matched = target_df[target_df[request.target_column].astype(str) == lookup_val]
            elif request.match_type == "contains":
                matched = target_df[target_df[request.target_column].astype(str).str.contains(lookup_val, case=False, na=False)]
            elif request.match_type == "regex":
                if request.lookup_value_pattern:
                    matched = target_df[target_df[request.target_column].astype(str).str.match(request.lookup_value_pattern, na=False)]
                else:
                    matched = target_df[target_df[request.target_column].astype(str).str.match(lookup_val, na=False)]
            
            if not matched.empty:
                result_val = matched.iloc[0][request.result_column]
                results.append({
                    "lookup_value": lookup_val,
                    "matched_value": str(matched.iloc[0][request.target_column]),
                    "result": str(result_val)
                })
                matches_count += 1
        
        return {
            "total_lookups": len(lookup_df),
            "matches_found": matches_count,
            "results": results[:100]  # Limit to first 100 for display
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/vlookup-write")
async def vlookup_write(request: VLookupRequest):
    """Perform VLOOKUP and write results to a new column"""
    try:
        print(f"Received request: {request}")
        file_path = os.path.join(UPLOAD_DIR, request.filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"File '{request.filename}' not found")
        
        workbook = openpyxl.load_workbook(file_path)
        
        # Read sheets as dataframes
        df_workbook = pd.read_excel(file_path, sheet_name=None)
        
        if request.lookup_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Lookup sheet '{request.lookup_sheet}' not found")
        if request.target_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Target sheet '{request.target_sheet}' not found")
        
        lookup_df = df_workbook[request.lookup_sheet]
        target_df = df_workbook[request.target_sheet]
        
        if request.lookup_column not in lookup_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.lookup_column}' not found in sheet '{request.lookup_sheet}'. Available columns: {list(lookup_df.columns)}")
        if request.target_column not in target_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.target_column}' not found in sheet '{request.target_sheet}'. Available columns: {list(target_df.columns)}")
        if request.result_column not in target_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.result_column}' not found in sheet '{request.target_sheet}'. Available columns: {list(target_df.columns)}")
        
        # Create new column name
        new_column_name = f"VLOOKUP_{request.result_column}"
        lookup_df[new_column_name] = None
        
        # Perform lookup for each row
        for idx, row in lookup_df.iterrows():
            lookup_val = str(row[request.lookup_column])
            
            if request.match_type == "exact":
                matched = target_df[target_df[request.target_column].astype(str) == lookup_val]
            elif request.match_type == "contains":
                matched = target_df[target_df[request.target_column].astype(str).str.contains(lookup_val, case=False, na=False)]
            elif request.match_type == "regex":
                if request.lookup_value_pattern:
                    matched = target_df[target_df[request.target_column].astype(str).str.match(request.lookup_value_pattern, na=False)]
                else:
                    matched = target_df[target_df[request.target_column].astype(str).str.match(lookup_val, na=False)]
            
            if not matched.empty:
                lookup_df.at[idx, new_column_name] = matched.iloc[0][request.result_column]
        
        # Write back to Excel
        output_filename = f"vlookup_{request.filename}"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in df_workbook.items():
                if sheet_name == request.lookup_sheet:
                    lookup_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        matches_count = lookup_df[new_column_name].notna().sum()
        
        return {
            "matches": int(matches_count),
            "total_rows": len(lookup_df),
            "new_column": new_column_name,
            "output_file": output_filename,
            "message": f"VLOOKUP completed: {matches_count} matches found out of {len(lookup_df)} rows"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in vlookup_write: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/merge-sheets")
async def merge_sheets(request: MergeSheetsRequest):
    """Merge two sheets based on a common key"""
    try:
        file_path = os.path.join(UPLOAD_DIR, request.filename)
        df_workbook = pd.read_excel(file_path, sheet_name=None)
        
        if request.left_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.left_sheet}' not found")
        if request.right_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.right_sheet}' not found")
        
        left_df = df_workbook[request.left_sheet]
        right_df = df_workbook[request.right_sheet]
        
        if request.left_key not in left_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.left_key}' not found in {request.left_sheet}")
        if request.right_key not in right_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.right_key}' not found in {request.right_sheet}")
        
        # Perform merge
        merged_df = pd.merge(
            left_df, 
            right_df, 
            left_on=request.left_key, 
            right_on=request.right_key, 
            how=request.merge_type,
            suffixes=('_left', '_right')
        )
        
        # Save to new file
        output_filename = f"merged_{request.filename}"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            merged_df.to_excel(writer, sheet_name='Merged_Data', index=False)
            for sheet_name, df in df_workbook.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return {
            "message": f"Sheets merged successfully: {len(merged_df)} rows",
            "rows": len(merged_df),
            "columns": len(merged_df.columns),
            "output_file": output_filename
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/xlookup")
async def xlookup(request: XLookupRequest):
    """Perform XLOOKUP with multiple return columns"""
    try:
        file_path = os.path.join(UPLOAD_DIR, request.filename)
        df_workbook = pd.read_excel(file_path, sheet_name=None)
        
        if request.lookup_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.lookup_sheet}' not found")
        if request.target_sheet not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.target_sheet}' not found")
        
        lookup_df = df_workbook[request.lookup_sheet]
        target_df = df_workbook[request.target_sheet]
        
        if request.lookup_column not in lookup_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.lookup_column}' not found in lookup sheet")
        if request.target_column not in target_df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{request.target_column}' not found in target sheet")
        
        for col in request.return_columns:
            if col not in target_df.columns:
                raise HTTPException(status_code=400, detail=f"Column '{col}' not found in target sheet")
        
        # Perform lookup for each return column
        for col in request.return_columns:
            new_col_name = f"XLOOKUP_{col}"
            lookup_df[new_col_name] = None
            
            for idx, row in lookup_df.iterrows():
                lookup_val = str(row[request.lookup_column])
                
                if request.match_type == "exact":
                    matched = target_df[target_df[request.target_column].astype(str) == lookup_val]
                elif request.match_type == "contains":
                    matched = target_df[target_df[request.target_column].astype(str).str.contains(lookup_val, case=False, na=False)]
                elif request.match_type == "regex":
                    matched = target_df[target_df[request.target_column].astype(str).str.match(lookup_val, na=False)]
                
                if not matched.empty:
                    lookup_df.at[idx, new_col_name] = matched.iloc[0][col]
        
        # Save to new file
        output_filename = f"xlookup_{request.filename}"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in df_workbook.items():
                if sheet_name == request.lookup_sheet:
                    lookup_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        matches_count = lookup_df[[f"XLOOKUP_{col}" for col in request.return_columns]].notna().any(axis=1).sum()
        
        return {
            "message": f"XLOOKUP completed: {matches_count} matches found",
            "matches": int(matches_count),
            "total_rows": len(lookup_df),
            "columns_added": len(request.return_columns),
            "output_file": output_filename
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/reconcile")
async def reconcile(request: ReconcileRequest):
    """Reconcile two sheets by comparing specific columns"""
    try:
        file_path = os.path.join(UPLOAD_DIR, request.filename)
        df_workbook = pd.read_excel(file_path, sheet_name=None)
        
        if request.sheet1 not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.sheet1}' not found")
        if request.sheet2 not in df_workbook:
            raise HTTPException(status_code=400, detail=f"Sheet '{request.sheet2}' not found")
        
        df1 = df_workbook[request.sheet1]
        df2 = df_workbook[request.sheet2]
        
        if request.key_column not in df1.columns:
            raise HTTPException(status_code=400, detail=f"Key column '{request.key_column}' not found in {request.sheet1}")
        if request.key_column not in df2.columns:
            raise HTTPException(status_code=400, detail=f"Key column '{request.key_column}' not found in {request.sheet2}")
        
        for col in request.compare_columns:
            if col not in df1.columns:
                raise HTTPException(status_code=400, detail=f"Column '{col}' not found in {request.sheet1}")
            if col not in df2.columns:
                raise HTTPException(status_code=400, detail=f"Column '{col}' not found in {request.sheet2}")
        
        # Merge on key column
        merged = pd.merge(df1, df2, on=request.key_column, how='outer', suffixes=('_sheet1', '_sheet2'), indicator=True)
        
        # Create reconciliation report
        reconcile_data = []
        
        for idx, row in merged.iterrows():
            key_value = row[request.key_column]
            
            if row['_merge'] == 'left_only':
                reconcile_data.append({
                    request.key_column: key_value,
                    'Status': 'Only in ' + request.sheet1,
                    'Difference': 'Missing in ' + request.sheet2
                })
            elif row['_merge'] == 'right_only':
                reconcile_data.append({
                    request.key_column: key_value,
                    'Status': 'Only in ' + request.sheet2,
                    'Difference': 'Missing in ' + request.sheet1
                })
            else:
                # Compare columns
                differences = []
                for col in request.compare_columns:
                    val1 = str(row[f'{col}_sheet1']) if pd.notna(row[f'{col}_sheet1']) else ''
                    val2 = str(row[f'{col}_sheet2']) if pd.notna(row[f'{col}_sheet2']) else ''
                    if val1 != val2:
                        differences.append(f"{col}: '{val1}' vs '{val2}'")
                
                if differences:
                    reconcile_data.append({
                        request.key_column: key_value,
                        'Status': 'Mismatch',
                        'Difference': '; '.join(differences)
                    })
        
        reconcile_df = pd.DataFrame(reconcile_data)
        
        # Save to new file
        output_filename = f"reconcile_{request.filename}"
        output_path = os.path.join(OUTPUT_DIR, output_filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            reconcile_df.to_excel(writer, sheet_name='Reconciliation_Report', index=False)
            merged.to_excel(writer, sheet_name='Merged_Comparison', index=False)
            for sheet_name, df in df_workbook.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        only_in_sheet1 = len([x for x in reconcile_data if 'Only in ' + request.sheet1 in x.get('Status', '')])
        only_in_sheet2 = len([x for x in reconcile_data if 'Only in ' + request.sheet2 in x.get('Status', '')])
        mismatches = len([x for x in reconcile_data if x.get('Status') == 'Mismatch'])
        
        return {
            "message": "Reconciliation completed",
            "total_differences": len(reconcile_df),
            "only_in_sheet1": only_in_sheet1,
            "only_in_sheet2": only_in_sheet2,
            "mismatches": mismatches,
            "output_file": output_filename
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/download/{filename}")
async def download_file(filename: str):
    """Download modified file"""
    file_path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="File not found")
    return FileResponse(file_path, filename=filename)

# Serve React App (must be last)
@app.get("/{full_path:path}")
async def serve_react_app(full_path: str):
    """Serve React app for all non-API routes"""
    if FRONTEND_BUILD.exists():
        file_path = FRONTEND_BUILD / full_path
        if file_path.is_file():
            return FileResponse(file_path)
        # If file doesn't exist, serve index.html (for React Router)
        return FileResponse(FRONTEND_BUILD / "index.html")
    return {"message": "Frontend not built. Run 'npm run build' in frontend directory"}

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
